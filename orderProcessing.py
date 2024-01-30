import requests
import pandas as pd
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from email_module import create_and_draft_email

current_date = datetime.now()
# Create date format for orders
year = current_date.strftime('%Y')
month_number = current_date.strftime('%m')
month_name = current_date.strftime('%B')
day = current_date.strftime('%d')

path_prefix = f'./{year}/{month_number} - {month_name}/{month_number}-{day}-{year}/'


class VendorFormatter:
    def __init__(self, location_id, output_filename):
        self.location_id = location_id
        self.output_filename = output_filename
        self.columns_order = []
        self.mapping = {}
        self.default_values = {}
        self.location_id_file = f"lastRun_{location_id}.txt"
        

    def set_defaults(self, default_values):
        if default_values is not None:
            self.default_values = default_values
        else:
            self.default_values = {}

    def set_mapping(self, columns_order, mapping):
        self.columns_order = columns_order
        self.mapping = mapping

    def extract_nested_field(self, node, field_path):
        """
        Extract a nested field from the GraphQL response node.
        Example: extract_nested_field(order_node, "shippingAddress.name")
        """
        fields = field_path.split(".")
        value = node
        for field in fields:
            if field.isdigit():  # Check if the field is a numeric index (indicating an array)
                field = int(field)
                if isinstance(value, list) and 0 <= field < len(value):
                    value = value[field]
                else:
                    return None
            elif field in value:
                value = value[field]
            else:
                return None  # Handle missing fields gracefully
            
        return value

    def map_vendor_data(self, order_node):
        all_line_items_data = []

        for line_item_node in order_node.get("lineItems", {}).get("nodes", []):
            if line_item_node.get("requiresShipping", True):
                vendor_data = {}
                for column_name, details in self.mapping.items():
                    if details["level"] == "order":
                        vendor_data[column_name] = self.extract_nested_field(order_node, details["path"])
                    elif details["level"] == "line_item":
                        vendor_data[column_name] = line_item_node.get(details["path"])
                    elif details["level"] == "static":
                        vendor_data[column_name] = details["value"]

                if "order" in order_node:
                    order_data = order_node["order"]
                    if "id" in order_data:
                        vendor_data["PO Number"] = order_data["id"].split("/")[-1]
                        vendor_data["id"] = order_data["id"].split("/")[-1]

                    if "customer" in order_data and "id" in order_data["customer"]:
                        vendor_data["Customer ID"] = order_data["customer"]["id"].split("/")[-1]

                # Set default values for columns that need them
                for column_name, default_value in self.default_values.items():
                    vendor_data.setdefault(column_name, default_value)

                all_line_items_data.append(vendor_data)

        return all_line_items_data




    
    def run_query_and_format(self):
        # Your Shopify GraphQL endpoint
        graphql_endpoint = os.getenv("ENDPOINT_URL")

        # Your Shopify access token
        access_token = os.getenv("ACCESS_TOKEN")

        # GraphQL query template with a variable
        graphql_query_template = """
            query FulfillmentOrders($locationId: String!, $after: String) {
                fulfillmentOrders(
                    first: 5
                    query: $locationId
                    includeClosed: true
                    after: $after
                ) {
                    nodes {
                        assignedLocation {
                            name
                        }
                        order {
                            id
                            name
                            billingAddress {
                                address1
                                address2
                                company
                                city
                                countryCode
                                firstName
                                lastName
                                name
                                phone
                                provinceCode
                                zip
                            }
                            customer {
                                id
                                email
                            }
                            shippingAddress {
                                address1
                                address2
                                company
                                city
                                countryCode
                                firstName
                                lastName
                                name
                                phone
                                provinceCode
                                zip
                            }
                            shippingLine {
                                title
                            }
                        }
                        status
                        lineItems(first: 25) {
                            pageInfo {
                                hasNextPage
                            }
                            edges {
                                cursor
                            }
                            nodes {
                                sku
                                totalQuantity
                                productTitle
                                requiresShipping
                            }
                        }
                        createdAt
                    }
                    pageInfo {
                        endCursor
                        hasNextPage
                    }
                    edges {
                        cursor
                    }
                }
            }
        """

        all_flattened_data = []
        cursor = None

        while True:
            if os.path.isfile(self.location_id_file):
                with open(self.location_id_file, "r") as file:
                    cursor = file.read().strip()

            response = requests.post(
                graphql_endpoint,
                json={"query": graphql_query_template, "variables": {"locationId": f"assigned_location_id:{self.location_id}", "after": cursor}},
                headers={"Content-Type": "application/json", "X-Shopify-Access-Token": access_token},
            )
            data = response.json()

            for order_node in data["data"]["fulfillmentOrders"]["nodes"]:
                line_items_data = self.map_vendor_data(order_node)
                all_flattened_data.extend(line_items_data)

            # Update cursor for the next page
            has_next_page = data["data"]["fulfillmentOrders"]["pageInfo"]["hasNextPage"]
            cursor = data["data"]["fulfillmentOrders"]["pageInfo"]["endCursor"]
            if has_next_page:                
                with open(self.location_id_file, "w") as file:
                    file.write(cursor)
            else:
                break

        if cursor is not None:
            with open(self.location_id_file, "w") as file:
                file.write(cursor)

        if all_flattened_data:
            df = pd.DataFrame(all_flattened_data, columns=self.columns_order)
            xlsx_file_path = f"{path_prefix}{self.output_filename}.xlsx"

            # Create directory structure if it doesn't already exist
            os.makedirs(os.path.dirname(xlsx_file_path), exist_ok=True)

            df.to_excel(xlsx_file_path, index=False)

            # Auto-adjust columns' width
            workbook = load_workbook(xlsx_file_path)
            worksheet = workbook.active

            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells) + 2
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

            workbook.save(xlsx_file_path)
            #print(f"XLSX file has been created: {xlsx_file_path}")
        else:
            print("No data to include in the DataFrame.")




# Load environment variables from .env
load_dotenv()

# List of vendors with their formatting details
vendors = [
    {
        "location_id": "00000000000",
        "vendor_name": "vendor1",
        "output_filename": "filename1",
        "email_addresses": [
            "example@example.com",
        ],
        "columns_order": ["Order ID", "Ship-to Name", "Ship-to Address 1", "Ship-to Address 2", "Ship-to City",
                          "Ship-to State", "Ship-to Zip", "", "SKU Number", "Quantity"],
        "mapping": {
            "Order ID": {"path": "order.name", "level": "order"},
            "Ship-to Name":  {"path": "order.shippingAddress.name", "level": "order"},
            "Ship-to Address 1":  {"path": "order.shippingAddress.address1", "level": "order"},
            "Ship-to Address 2":  {"path": "order.shippingAddress.address2", "level": "order"},
            "Ship-to City":  {"path": "order.shippingAddress.city", "level": "order"},
            "Ship-to State":  {"path": "order.shippingAddress.provinceCode", "level": "order"},
            "Ship-to Zip":  {"path": "order.shippingAddress.zip", "level": "order"},
            "": {"path": "static", "level": "static", "value": "Home"},
            "SKU Number":  {"path": "sku", "level": "line_item"},
            "Quantity":  {"path": "totalQuantity", "level": "line_item"},
        },
        "defaults": {
            "": "Home"
        }
    },
    {
        "location_id":"00000000000",
        "vendor_name": "vendor2",
        "output_filename": "filename2",
        "email_addresses": [
            "example@example.com"
        ],
        "columns_order": ["Name", "Email", "Created At", "Shipping Method", "Lineitem quantity",
            "Lineitem Name", "Lineitem sku", "Billing Name", "Billing Address1", "Billing Address2",
            "Billing Company", "Billing City", "Billing Zip", "Billing Province", "Billing Country",
            "Billing Phone", "Shipping Name", "Shipping Address1", "Shipping Address2", "Shipping Company",
            "Shipping City", "Shipping Zip", "Shipping Province", "Shipping Country", "Shipping Phone", "Id"],
        "mapping": {
            "Name": {"path": "order.name", "level": "order"},
            "Email": {"path": "order.customer.email", "level": "order"},
            "Created At": {"path": "createdAt", "level": "order"},
            "Shipping Method": {"path": "order.shippingLine.title", "level": "order"},
            "Lineitem quantity": {"path": "totalQuantity", "level": "line_item"},
            "Lineitem Name": {"path": "productTitle", "level": "line_item"},
            "Lineitem sku": {"path": "sku", "level": "line_item"},
            "Billing Name": {"path": "order.billingAddress.name", "level": "order"},
            "Billing Address1": {"path": "order.billingAddress.address1", "level": "order"},
            "Billing Address2": {"path": "order.billingAddress.address2", "level": "order"},
            "Billing Company": {"path": "order.billingAddress.company", "level": "order"},
            "Billing City": {"path": "order.billingAddress.city", "level": "order"},
            "Billing Zip": {"path": "order.billingAddress.zip", "level": "order"},
            "Billing Province": {"path": "order.billingAddress.provinceCode", "level": "order"},
            "Billing Country": {"path": "order.billingAddress.countryCode", "level": "order"},
            "Billing Phone": {"path": "order.billingAddress.phone", "level": "order"},
            "Shipping Name": {"path": "order.shippingAddress.name", "level": "order"},
            "Shipping Address1": {"path": "order.shippingAddress.address1", "level": "order"},
            "Shipping Address2": {"path": "order.shippingAddress.address2", "level": "order"},
            "Shipping Company": {"path": "order.shippingAddress.company", "level": "order"},
            "Shipping City": {"path": "order.shippingAddress.city", "level": "order"},
            "Shipping Zip": {"path": "order.shippingAddress.zip", "level": "order"},
            "Shipping Province": {"path": "order.shippingAddress.provinceCode", "level": "order"},
            "Shipping Country": {"path": "order.shippingAddress.countryCode", "level": "order"},
            "Shipping Phone": {"path": "order.shippingAddress.phone", "level": "order"},
            "Id": {"path": "order.id", "level": "order"},
        },
        "defaults": {

        }
    },
    {
        "location_id": "00000000000",
        "vendor_name": "vendor3",
        "output_filename": "filename3",
        "email_addresses": [
            "example@example.com",
            "example2@example.com",
        ],
        "columns_order": ["Order Number", "PO Number", "Order Date", "Customer ID", "Billing Full Name",
                          "Billing First Name", "Billing Last Name", "Billing Street 1", "Billing Street 2", "Billing City",
                          "Billing State", "Billing ZipCode", "Billing Email", "Billing Phone",
                          "Billing Country", "Shipping Full Name", "Shipping First Name", "Shipping Last Name",
                          "Shipping Street 1", "Shipping Street 2", "Shipping City", "Shipping State",
                          "Shipping ZipCode", "Shipping Email", "Shipping Phone", "Shipping Country",
                          "Shipping Charge", "Sales Tax", "Shipping Carrier", "Shipping Method", "Ship By Date",
                          "Shipping Reference 1", "Shipping Reference 2", "3rd Party Shipping Account Number",
                          "3rd Party Shipping Name", "3rd Party Shipping Address", "3rd Party Shipping City",
                          "3rd Party Shipping State", "3rd Party Shipping ZipCode", "3rd Party Shipping Country",
                          "Lineitem sku", "Item Description", "Item Quantity", "Item Price", "Notes"],
        "mapping": {
            "Order Number": {"path": "order.name", "level": "order"},
            "PO Number": {"path": "order.id", "level": "order"},
            "Order Date": {"path": "createdAt", "level": "order"},
            "Customer ID": {"path": "order.customer.id", "level": "order"},
            "Billing Full Name": {"path": "order.billingAddress.name", "level": "order"},
            "Billing First Name": {"path": "order.billingAddress.firstName", "level": "order"},
            "Billing Last Name": {"path": "order.billingAddress.lastName", "level": "order"},
            "Billing Street 1": {"path": "order.billingAddress.address1", "level": "order"},
            "Billing Street 2": {"path": "order.billingAddress.address2", "level": "order"},
            "Billing City": {"path": "order.billingAddress.city", "level": "order"},
            "Billing State": {"path": "order.billingAddress.provinceCode", "level": "order"},
            "Billing ZipCode": {"path": "order.billingAddress.zip", "level": "order"},
            "Billing Email": {"path": "order.customer.email", "level": "order"},
            "Billing Phone": {"path": "order.billingAddress.phone", "level": "order"},
            "Billing Country": {"path": "order.billingAddress.countryCode", "level": "order"},
            "Shipping Full Name": {"path": "order.shippingAddress.name", "level": "order"},
            "Shipping First Name": {"path": "order.shippingAddress.firstName", "level": "order"},
            "Shipping Last Name": {"path": "order.shippingAddress.lastName", "level": "order"},
            "Shipping Street 1": {"path": "order.shippingAddress.address1", "level": "order"},
            "Shipping Street 2": {"path": "order.shippingAddress.address2", "level": "order"},
            "Shipping City": {"path": "order.shippingAddress.city", "level": "order"},
            "Shipping State": {"path": "order.shippingAddress.provinceCode", "level": "order"},
            "Shipping ZipCode": {"path": "order.shippingAddress.zip", "level": "order"},
            "Shipping Email": {"path": "order.customer.email", "level": "order"},
            "Shipping Phone": {"path": "order.shippingAddress.phone", "level": "order"},
            "Shipping Country": {"path": "order.shippingAddress.countryCode", "level": "order"},
            "Shipping Charge": {"path": "static", "level": "static", "value": ""},
            "Sales Tax": {"path": "static", "level": "static", "value": ""},
            "Shipping Carrier": {"path": "static", "level": "static", "value": "FedEx"},
            "Shipping Method": {"path": "order.shippingLine.title", "level": "order"},
            "Ship By Date": {"path": "static", "level": "static", "value": ""},
            "Shipping Reference 1": {"path": "static", "level": "static", "value": ""},
            "Shipping Reference 2": {"path": "static", "level": "static", "value": ""},
            "3rd Party Shipping Account Number": {"path": "static", "level": "static", "value": "000000000"},
            "3rd Party Shipping Name": {"path": "static", "level": "static", "value": ""},
            "3rd Party Shipping Address": {"path": "static", "level": "static", "value": ""},
            "3rd Party Shipping City": {"path": "static", "level": "static", "value": ""},
            "3rd Party Shipping State": {"path": "static", "level": "static", "value": ""},
            "3rd Party Shipping ZipCode": {"path": "static", "level": "static", "value": ""},
            "3rd Party Shipping Country": {"path": "static", "level": "static", "value": ""},
            "Lineitem sku": {"path": "sku", "level": "line_item"},
            "Item Description": {"path": "productTitle", "level": "line_item"},
            "Item Quantity": {"path": "totalQuantity", "level": "line_item"},
            "Item Price": {"path": "static", "level": "static", "value": ""},
            "Notes": {"path": "static", "level": "static", "value": ""}
        },
        "defaults": {
            "Shipping Charge": "",
            "Sales Tax": "",
            "Shipping Carrier": "FedEx",
            "Ship By Date": "",
            "Shipping Reference 1": "",
            "Shipping Reference 2": "",
            "3rd Party Shipping Account Number": "000000000",
            "3rd Party Shipping Name": "",
            "3rd Party Shipping Address": "",
            "3rd Party Shipping City": "",
            "3rd Party Shipping State": "",
            "3rd Party Shipping ZipCode": "",
            "3rd Party Shipping Country": "",
            "Item Price": "",
            "Notes": ""
        }
    }
    # Add similar entries for other vendors
]

# Loop through vendors
for vendor_details in vendors:
    vendor_formatter = VendorFormatter(vendor_details["location_id"], vendor_details["output_filename"])
    vendor_formatter.set_mapping(vendor_details["columns_order"], vendor_details["mapping"])
    vendor_formatter.set_defaults(vendor_details.get("defaults",{}))
    vendor_formatter.run_query_and_format()

    # Email the generated file(s)
    vendor_name = f"{vendor_details['vendor_name']}"
    email_subject = f"Orders {month_number}-{day}"
    email_recipient = vendor_details.get("email_addresses", [])
    attachment_path = f"{path_prefix}{vendor_details['output_filename']}.xlsx"
    cc_emails = ["cc@example.com", "cc2@example.com"]
    email_body = "Place your email text here, using \n for line breaks"
    email_signature_html = """
If you have an HTML/formatted/styled email signature, you can paste the code here.
"""

    create_and_draft_email(email_recipient, email_subject, email_body, email_signature_html, attachment_path, cc_emails, vendor_name)
input("Finished processing orders, press Enter to exit...")    
